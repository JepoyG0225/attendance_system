"""
Monthly attendance report generation.
Produces Excel (.xlsx) and PDF (.pdf) summaries per student/section/grade.
"""
from __future__ import annotations

import io
from calendar import monthrange
from datetime import date, timedelta
from typing import Optional

from sqlalchemy.orm import Session

from database import Student, Attendance, Section, Grade, AttendanceStatus
from config import SCHOOL_NAME


# ── Data aggregation ─────────────────────────────────────────────────────────

def _month_dates(year: int, month: int) -> list[date]:
    _, last_day = monthrange(year, month)
    return [date(year, month, d) for d in range(1, last_day + 1)]


def _aggregate(year: int, month: int, db: Session,
               section_id: Optional[int] = None,
               grade_id: Optional[int] = None) -> dict:
    """
    Build the per-student rollup for the month.
    Returns a dict with: meta, headers, students[], totals
    """
    from attendance import is_school_day  # local import to avoid circular

    days = _month_dates(year, month)
    school_days = [d for d in days if is_school_day(d, db)[0]]
    total_school_days = len(school_days)

    q = db.query(Student).filter(Student.is_active == True)
    if section_id:
        q = q.filter(Student.section_id == section_id)
    elif grade_id:
        section_ids = [s.id for s in db.query(Section).filter(Section.grade_id == grade_id).all()]
        q = q.filter(Student.section_id.in_(section_ids or [-1]))
    students = q.order_by(Student.full_name).all()

    rows = []
    totals = {"present": 0, "late": 0, "absent": 0, "excused": 0, "no_record": 0}

    for s in students:
        records = (
            db.query(Attendance)
            .filter(
                Attendance.student_id == s.id,
                Attendance.date >= days[0],
                Attendance.date <= days[-1],
            )
            .all()
        )
        by_date = {r.date: r for r in records}

        counts = {"present": 0, "late": 0, "absent": 0, "excused": 0, "no_record": 0}
        for d in school_days:
            rec = by_date.get(d)
            if rec is None:
                counts["no_record"] += 1
            else:
                counts[rec.status.value] += 1

        for k, v in counts.items():
            totals[k] += v

        attendance_rate = 0.0
        accounted = counts["present"] + counts["late"]
        if total_school_days:
            attendance_rate = round((accounted / total_school_days) * 100, 1)

        rows.append({
            "id":         s.id,
            "name":       s.full_name,
            "grade":      s.grade_name,
            "section":    s.section_name,
            "rfid":       s.rfid_uid,
            "parent":     s.parent_name,
            "phone":      s.parent_phone,
            "present":    counts["present"],
            "late":       counts["late"],
            "absent":     counts["absent"],
            "excused":    counts["excused"],
            "no_record":  counts["no_record"],
            "rate":       attendance_rate,
        })

    return {
        "school":             SCHOOL_NAME,
        "year":               year,
        "month":              month,
        "month_name":         date(year, month, 1).strftime("%B %Y"),
        "school_days":        total_school_days,
        "calendar_days":      len(days),
        "students":           rows,
        "totals":             totals,
        "filter_section_id":  section_id,
        "filter_grade_id":    grade_id,
    }


# ── Excel (xlsx) ─────────────────────────────────────────────────────────────

def build_xlsx(year: int, month: int, db: Session,
               section_id: Optional[int] = None,
               grade_id: Optional[int] = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    data = _aggregate(year, month, db, section_id=section_id, grade_id=grade_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    title_font  = Font(bold=True, size=14, color="FFFFFF")
    title_fill  = PatternFill("solid", fgColor="1D4ED8")
    sub_font    = Font(bold=True, size=10, color="475569")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="334155")
    border      = Border(*(Side(style="thin", color="CBD5E1") for _ in range(4)))
    center      = Alignment(horizontal="center", vertical="center")

    # ── Title rows ────────────────────────────────────────────────────────────
    ws["A1"] = data["school"]
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws.merge_cells("A1:L1")
    ws.row_dimensions[1].height = 26

    ws["A2"] = f"Monthly Attendance — {data['month_name']}"
    ws["A2"].font = Font(bold=True, size=12)
    ws.merge_cells("A2:L2")

    filter_label = ""
    if section_id:
        sec = db.query(Section).filter(Section.id == section_id).first()
        if sec: filter_label = f"Section: {sec.grade.name if sec.grade else ''} {sec.name}"
    elif grade_id:
        g = db.query(Grade).filter(Grade.id == grade_id).first()
        if g: filter_label = f"Grade: {g.name}"
    ws["A3"] = (filter_label + "   ·   ") if filter_label else ""
    ws["A3"].value = (filter_label + "   ·   " if filter_label else "") + \
                     f"School days in month: {data['school_days']} of {data['calendar_days']}"
    ws["A3"].font = sub_font
    ws.merge_cells("A3:L3")

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["#", "Name", "Grade", "Section", "RFID", "Parent", "Phone",
               "Present", "Late", "Absent", "Excused", "Rate %"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, s in enumerate(data["students"], start=1):
        row_idx = 5 + i
        values = [i, s["name"], s["grade"], s["section"], s["rfid"],
                  s["parent"], s["phone"],
                  s["present"], s["late"], s["absent"], s["excused"], s["rate"]]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col, value=v)
            c.border = border
            if col >= 8 or col == 1:
                c.alignment = center
        # color the rate cell
        rate = s["rate"]
        rate_cell = ws.cell(row=row_idx, column=12)
        if rate >= 90:
            rate_cell.fill = PatternFill("solid", fgColor="DCFCE7")
        elif rate >= 75:
            rate_cell.fill = PatternFill("solid", fgColor="FEF3C7")
        else:
            rate_cell.fill = PatternFill("solid", fgColor="FEE2E2")

    # ── Totals row ────────────────────────────────────────────────────────────
    total_row = 5 + len(data["students"]) + 1
    ws.cell(row=total_row, column=1, value="TOTALS").font = Font(bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
    for col, k in enumerate(["present", "late", "absent", "excused"], start=8):
        c = ws.cell(row=total_row, column=col, value=data["totals"][k])
        c.font = Font(bold=True)
        c.alignment = center

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [4, 26, 12, 16, 14, 22, 16, 9, 8, 9, 10, 9]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A6"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ──────────────────────────────────────────────────────────────────────

def build_pdf(year: int, month: int, db: Session,
              section_id: Optional[int] = None,
              grade_id: Optional[int] = None) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    data = _aggregate(year, month, db, section_id=section_id, grade_id=grade_id)
    styles = getSampleStyleSheet()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=12*mm, bottomMargin=12*mm,
                            title=f"{data['school']} - {data['month_name']}")
    story = []

    title_style = ParagraphStyle("Title", parent=styles["Heading1"],
                                 fontSize=16, alignment=1, textColor=colors.HexColor("#0F172A"))
    sub_style   = ParagraphStyle("Sub",   parent=styles["Normal"],
                                 fontSize=10, alignment=1, textColor=colors.HexColor("#475569"))

    story.append(Paragraph(data["school"], title_style))
    story.append(Paragraph(f"Monthly Attendance Report — {data['month_name']}", sub_style))

    filter_label = ""
    if section_id:
        sec = db.query(Section).filter(Section.id == section_id).first()
        if sec: filter_label = f"Section: {sec.grade.name if sec.grade else ''} {sec.name}"
    elif grade_id:
        g = db.query(Grade).filter(Grade.id == grade_id).first()
        if g: filter_label = f"Grade: {g.name}"
    meta_line = (filter_label + "   ·   " if filter_label else "") + \
                f"School days: {data['school_days']} of {data['calendar_days']}"
    story.append(Paragraph(meta_line, sub_style))
    story.append(Spacer(1, 8))

    headers = ["#", "Name", "Grade", "Section", "Parent", "Phone",
               "Present", "Late", "Absent", "Excused", "Rate %"]
    rows = [headers]
    for i, s in enumerate(data["students"], start=1):
        rows.append([
            i, s["name"], s["grade"], s["section"],
            s["parent"], s["phone"],
            s["present"], s["late"], s["absent"], s["excused"], f"{s['rate']:.1f}",
        ])
    rows.append(["", "TOTALS", "", "", "", "",
                 data["totals"]["present"], data["totals"]["late"],
                 data["totals"]["absent"], data["totals"]["excused"], ""])

    table = Table(rows, repeatRows=1)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("ALIGN",      (1, 1), (5, -1), "LEFT"),
        ("GRID",       (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
        ("FONTNAME",   (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F8FAFC")]),
    ])
    # color-tint the rate column
    for i, s in enumerate(data["students"], start=1):
        if s["rate"] >= 90:   bg = colors.HexColor("#DCFCE7")
        elif s["rate"] >= 75: bg = colors.HexColor("#FEF3C7")
        else:                 bg = colors.HexColor("#FEE2E2")
        style.add("BACKGROUND", (10, i), (10, i), bg)
    table.setStyle(style)

    story.append(table)
    doc.build(story)
    return buf.getvalue()
